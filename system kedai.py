from openpyxl import load_workbook
from datetime import datetime

# Buka fail Excel
workbook = load_workbook(filename='menu.xlsx', data_only=True)

# Pilih sheet
sheet_makanan = workbook['menu_makanan']
sheet_minuman = workbook['menu_minuman']
sheet_desert = workbook['menu_desert']

# Fungsi untuk baca menu dan harga dari sheet
def baca_menu(sheet):
    menu = []
    harga = []
    harga_agen_o = []
    harga_agen_n = []
    stok = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 untuk skip header
        nama = row[0] if len(row) > 0 and row[0] is not None else "Tanpa Nama"
        harga_item = row[1] if len(row) > 1 and row[1] is not None else 0.0
        h_agen_o = row[2] if len(row) > 2 and row[2] is not None else 0.0
        h_agen_n = row[3] if len(row) > 3 and row[3] is not None else 0.0
        stok_item = row[4] if len(row) > 4 and row[4] is not None else 0

        menu.append(nama)
        harga.append(harga_item)
        harga_agen_o.append(h_agen_o)
        harga_agen_n.append(h_agen_n)
        stok.append(stok_item)
    return menu, harga, harga_agen_o, harga_agen_n, stok


# Dapatkan data menu
menu_makanan, harga_retail_makanan, harga_agen_o_makanan, harga_agen_n_makanan, stok_makanan = baca_menu(sheet_makanan)
menu_minuman, harga_retail_minuman, harga_agen_o_minuman, harga_agen_n_minuman, stok_minuman = baca_menu(sheet_minuman)
menu_desert, harga_retail_desert, harga_agen_o_desert, harga_agen_n_desert, stok_desert = baca_menu(sheet_desert)

pesanan_makanan = []
pesanan_minuman = []
pesanan_desert = []

# Minta jenis pelanggan
print("Adakah anda pelanggan biasa atau agen?")
print("1. Pelanggan Biasa (Harga Retail)")
print("2. Agen O (Harga Agen O)")
print("3. Agen N (Harga Agen N)")
jenis_pelanggan = input("Sila pilih nombor (1/2/3): ")

# Fungsi untuk dapatkan harga ikut jenis pelanggan, handle None harga agen
def dapatkan_harga(index, kategori, jenis, pelanggan):
    # kategori: 'makanan', 'minuman', 'desert'
    if kategori == 'makanan':
        if pelanggan == '1':  # retail
            return harga_retail_makanan[index]
        elif pelanggan == '2':  # agen o
            return harga_agen_o_makanan[index] if harga_agen_o_makanan[index] is not None else harga_retail_makanan[index]
        elif pelanggan == '3':  # agen n
            return harga_agen_n_makanan[index] if harga_agen_n_makanan[index] is not None else harga_retail_makanan[index]
    elif kategori == 'minuman':
        if pelanggan == '1':
            return harga_retail_minuman[index]
        elif pelanggan == '2':
            return harga_agen_o_minuman[index] if harga_agen_o_minuman[index] is not None else harga_retail_minuman[index]
        elif pelanggan == '3':
            return harga_agen_n_minuman[index] if harga_agen_n_minuman[index] is not None else harga_retail_minuman[index]
    elif kategori == 'desert':
        if pelanggan == '1':
            return harga_retail_desert[index]
        elif pelanggan == '2':
            return harga_agen_o_desert[index] if harga_agen_o_desert[index] is not None else harga_retail_desert[index]
        elif pelanggan == '3':
            return harga_agen_n_desert[index] if harga_agen_n_desert[index] is not None else harga_retail_desert[index]
    return 0

# Papar menu lengkap dengan harga ikut retail sahaja (biasa)
def papar_menu():
    print("\n📋 Menu Makanan :")
    for i in range(len(menu_makanan)):
        print(f"{i + 1}. {menu_makanan[i]} - RM{harga_retail_makanan[i]:.2f}")
    print("\n📋 Menu Minuman :")
    offset = len(menu_makanan)
    for i in range(len(menu_minuman)):
        print(f"{offset + i + 1}. {menu_minuman[i]} - RM{harga_retail_minuman[i]:.2f}")
    print("\n📋 Menu Desert :")
    offset = len(menu_makanan) + len(menu_minuman)
    for i in range(len(menu_desert)):
        print(f"{offset + i + 1}. {menu_desert[i]} - RM{harga_retail_desert[i]:.2f}")

def tambah_pesanan(pilihan):
    if pilihan.isdigit():
        nombor = int(pilihan)
        if 1 <= nombor <= len(menu_makanan):
            index = nombor - 1
            if stok_makanan[index] > 0:
                pesanan_makanan.append(index)
                stok_makanan[index] -= 1
                print(f"✔️ {menu_makanan[index]} telah ditambah. Stok tinggal: {stok_makanan[index]}")
            else:
                print(f"❌ {menu_makanan[index]} habis stok!")
        elif len(menu_makanan) < nombor <= len(menu_makanan) + len(menu_minuman):
            index = nombor - len(menu_makanan) - 1
            if stok_minuman[index] > 0:
                pesanan_minuman.append(index)
                stok_minuman[index] -= 1
                print(f"✔️ {menu_minuman[index]} telah ditambah. Stok tinggal: {stok_minuman[index]}")
            else:
                print(f"❌ {menu_minuman[index]} habis stok!")
        elif len(menu_makanan) + len(menu_minuman) < nombor <= len(menu_makanan) + len(menu_minuman) + len(menu_desert):
            index = nombor - len(menu_makanan) - len(menu_minuman) - 1
            if stok_desert[index] > 0:
                pesanan_desert.append(index)
                stok_desert[index] -= 1
                print(f"✔️ {menu_desert[index]} telah ditambah. Stok tinggal: {stok_desert[index]}")
            else:
                print(f"❌ {menu_desert[index]} habis stok!")
        else:
            print("❌ Nombor tidak ada dalam menu.")
    else:
        print("❌ Sila taip nombor atau 'keluar'.")

def papar_dan_simpan_resit(nama_pelanggan):
    now = datetime.now()
    tarikh_masa = now.strftime("%d/%m/%Y %H:%M:%S")
    no_resit = now.strftime("NANI%Y%m%d%H%M%S")

    resit = logo
    resit += f"🧍‍♀️ Pelanggan: {nama_pelanggan}\n"
    resit += f"📄 No Resit : {no_resit}\n"
    resit += f"🕒 Tarikh & Masa: {tarikh_masa}\n"
    resit += "------------------------------\n"

    jumlah = 0

    # Makanan
    resit += "\n🍴 Makanan:\n"
    print("\n🍴 Makanan:")
    kiraan_makanan = {}
    for i in pesanan_makanan:
        kiraan_makanan[i] = kiraan_makanan.get(i, 0) + 1

    for i, kuantiti in kiraan_makanan.items():
        harga = dapatkan_harga(i, 'makanan', i, jenis_pelanggan)
        total = harga * kuantiti
        jumlah += total
        item_line = f"{menu_makanan[i]} x{kuantiti} - RM{total:.2f}"
        print(item_line)
        resit += item_line + "\n"

    # Minuman
    resit += "\n🥤 Minuman:\n"
    print("\n🥤 Minuman:")
    kiraan_minuman = {}
    for i in pesanan_minuman:
        kiraan_minuman[i] = kiraan_minuman.get(i, 0) + 1

    for i, kuantiti in kiraan_minuman.items():
        harga = dapatkan_harga(i, 'minuman', i, jenis_pelanggan)
        total = harga * kuantiti
        jumlah += total
        item_line = f"{menu_minuman[i]} x{kuantiti} - RM{total:.2f}"
        print(item_line)
        resit += item_line + "\n"

    # Desert
    resit += "\n🍨 Desert:\n"
    print("\n🍨 Desert:")
    kiraan_desert = {}
    for i in pesanan_desert:
        kiraan_desert[i] = kiraan_desert.get(i, 0) + 1

    for i, kuantiti in kiraan_desert.items():
        harga = dapatkan_harga(i, 'desert', i, jenis_pelanggan)
        total = harga * kuantiti
        jumlah += total
        item_line = f"{menu_desert[i]} x{kuantiti} - RM{total:.2f}"
        print(item_line)
        resit += item_line + "\n"

    resit += "------------------------------\n"
    resit += f"💰 Jumlah Perlu Dibayar: RM{jumlah:.2f}\n"
    resit += "Terima kasih kerana membeli! 🩵\n"

    print(f"\n💰 Jumlah Perlu Dibayar: RM{jumlah:.2f}")
    print(f"📄 Resit disimpan sebagai: resit_{no_resit}.txt")

    nama_fail = f"resit_{no_resit}.txt"
    with open(nama_fail, "w", encoding="utf-8") as f:
        f.write(resit)

def simpan_stok():
    for i, stok in enumerate(stok_makanan):
        sheet_makanan.cell(row=i+2, column=5).value = stok  # pastikan kolum stok ada betul
    for i, stok in enumerate(stok_minuman):
        sheet_minuman.cell(row=i+2, column=5).value = stok
    for i, stok in enumerate(stok_desert):
        sheet_desert.cell(row=i+2, column=5).value = stok
    workbook.save("menu.xlsx")

# Mula Program
logo = """
🏪🍽️ KEDAI FROZEN FAZLI 🍽️🏪
------------------------------
"""

print("🛍️ Selamat datang ke Kedai Frozen Fazli!")
nama_pelanggan = input("Masukkan nama anda: ")

papar_menu()

while True:
    pilihan = input("\nTaip nombor menu untuk pesan, atau 'keluar' untuk tamat: ")
    if pilihan.lower() == "keluar":
        break
    tambah_pesanan(pilihan)

if pesanan_makanan or pesanan_minuman or pesanan_desert:
    papar_dan_simpan_resit(nama_pelanggan)
    simpan_stok()
else:
    print("❗ Tiada pesanan dibuat. Jumpa lagi!")
